import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import telebot
from dotenv import load_dotenv
from openpyxl.formatting.rule import ColorScaleRule

# Load environment variables
load_dotenv()

# Get Telegram bot token and allowed user ID from .env
API_TOKEN = os.getenv('API_TOKEN')
ALLOWED_USER_ID = int(os.getenv('ALLOWED_USER_ID'))

# Initialize Telegram bot
bot = telebot.TeleBot(API_TOKEN)

# Function to fetch data from SQLite database
def fetch_data_from_db(db_path):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Fetching all data from the transactions table
    cursor.execute("SELECT * FROM transactions")
    rows = cursor.fetchall()
    
    conn.close()
    
    return rows

# Function to auto-resize columns based on content length
def auto_resize_columns(ws, data):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Function to add borders to all cells in the worksheet
def add_borders(ws):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

def generate_excel_report(data, report_file):
    wb = Workbook()

    headers = ["id", "type", "category", "amount", "description", "creation_date", "creation_time"]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Month", "Income", "Expense"])

    for col in ws_summary.columns:
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill

    add_borders(ws_summary)

    month_data = {}
    monthly_income_expense = {}

    for row in data:
        created_at = datetime.strptime(row[5], "%Y-%m-%d %H:%M:%S")
        month_str = created_at.strftime("%Y%m")

        if month_str not in month_data:
            month_data[month_str] = []
            monthly_income_expense[month_str] = {"income": 0, "expense": 0}

        month_data[month_str].append(row)

        amount = row[3]
        if row[1] == 'expense':
            monthly_income_expense[month_str]["expense"] += amount
        else:
            monthly_income_expense[month_str]["income"] += amount

    for month, totals in monthly_income_expense.items():
        ws_summary.append([month, totals["income"], totals["expense"]])

    ws_summary.auto_filter.ref = ws_summary.dimensions
    auto_resize_columns(ws_summary, data)

    # Apply conditional formatting
    last_row = ws_summary.max_row
    income_range = f"B2:B{last_row}"
    expense_range = f"C2:C{last_row}"

    green_yellow_red = ColorScaleRule(
        start_type='min', start_color='63BE7B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='F8696B'
    )

    red_yellow_green = ColorScaleRule(
        start_type='min', start_color='F8696B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'
    )

    ws_summary.conditional_formatting.add(income_range, red_yellow_green)
    ws_summary.conditional_formatting.add(expense_range, green_yellow_red)

    for month, month_rows in month_data.items():
        ws = wb.create_sheet(title=month)
        ws.append(headers)
        for col in ws.columns:
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill

        for row in month_rows:
            created_at = datetime.strptime(row[5], "%Y-%m-%d %H:%M:%S")
            new_row = list(row[:5]) + [created_at.strftime("%Y-%m-%d"), created_at.strftime("%H:%M:%S")]
            ws.append(new_row)

        ws.auto_filter.ref = ws.dimensions
        auto_resize_columns(ws, month_rows)
        add_borders(ws)

    wb.save(report_file)
# Function to send the file to the specified Telegram user
def send_report_to_telegram(report_file):
    with open(report_file, 'rb') as file:
        bot.send_document(ALLOWED_USER_ID, file)

# Main function to execute the script
def main():
    db_path = "app_kakeibo.db"
    report_file = "transactions_report.xlsx"
    
    # Fetch data from database
    data = fetch_data_from_db(db_path)
    
    # Generate Excel report
    generate_excel_report(data, report_file)

    # Send report to Telegram user
    send_report_to_telegram(report_file)

    # Delete the report file after sending
    os.remove(report_file)

if __name__ == "__main__":
    main()
